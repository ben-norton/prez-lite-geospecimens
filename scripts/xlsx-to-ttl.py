"""Generate prez-lite-conformant SKOS TTL files from sources/xlsx workbooks.

Reads each *.xlsx in sources/xlsx/, extracts the Concept Scheme, Concepts, and
Agents sheets, and writes a corresponding .ttl file to data/vocabs/.

Namespace bindings and property mappings follow sources/VocabularyAnnotations.xlsx.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX_DIR = ROOT / "sources" / "xlsx"
OUT_DIR = ROOT / "data" / "vocabs"

# Prefix bindings used in generated TTL (keep aligned with VocabularyAnnotations).
PREFIXES: list[tuple[str, str]] = [
    ("dcterms", "http://purl.org/dc/terms/"),
    ("isoroles", "https://linked.data.gov.au/def/data-roles/"),
    ("owl", "http://www.w3.org/2002/07/owl#"),
    ("prov", "http://www.w3.org/ns/prov#"),
    ("rdfs", "http://www.w3.org/2000/01/rdf-schema#"),
    ("reg", "http://purl.org/linked-data/registry#"),
    ("schema", "https://schema.org/"),
    ("skos", "http://www.w3.org/2004/02/skos/core#"),
    ("vann", "http://purl.org/vocab/vann/"),
    ("xsd", "http://www.w3.org/2001/XMLSchema#"),
]


def normalize_iri(iri: str) -> str:
    iri = iri.strip()
    if iri.startswith("http://vocab.geospecimens.org"):
        iri = "https://" + iri[len("http://") :]
    return iri


def parse_date(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, (datetime, date)):
        d = val.date() if isinstance(val, datetime) else val
        return d.isoformat()
    s = str(val).strip()
    if not s:
        return None
    # Accept YYYYMMDD numerics like 20260417.0
    m = re.match(r"^(\d{4})(\d{2})(\d{2})(?:\.0+)?$", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    return None


def iri_slug(iri: str) -> str:
    iri = iri.rstrip("/")
    tail = iri.rsplit("/", 1)[-1]
    return tail


def escape_string(s: str) -> str:
    s = str(s).replace("\\", "\\\\").replace('"', '\\"')
    return s


def ttl_string(s: str, lang: str | None = "en") -> str:
    s = str(s).strip()
    # Use triple-quoted form for multi-line strings
    if "\n" in s:
        body = s.replace("\\", "\\\\").replace('"""', '\\"\\"\\"')
        return f'"""{body}"""' + (f"@{lang}" if lang else "")
    body = escape_string(s)
    return f'"{body}"' + (f"@{lang}" if lang else "")


def ttl_plain(s: str) -> str:
    return f'"{escape_string(str(s).strip())}"'


def split_list(val: str | None) -> list[str]:
    if val is None:
        return []
    s = str(val).strip()
    if not s:
        return []
    parts = re.split(r"[,\n]+", s)
    return [p.strip() for p in parts if p and p.strip()]


def load_sheet_kv(ws) -> dict[str, list]:
    """Return a dict mapping each key (column A) to a list of its non-empty values."""
    out: dict[str, list] = {}
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        key = str(row[0]).strip()
        vals = [v for v in row[1:] if v not in (None, "") and str(v).strip() != ""]
        if key in out:
            out[key].extend(vals)
        else:
            out[key] = list(vals)
    return out


def load_concepts(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else None for h in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(c not in (None, "") for c in row):
            continue
        rec = {}
        for h, v in zip(headers, row):
            if h and v not in (None, ""):
                rec[h] = str(v).strip() if not isinstance(v, (int, float, datetime, date)) else v
        if rec.get("conceptIRI"):
            records.append(rec)
    return records


def load_agents(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else None for h in rows[0]]
    agents = []
    for row in rows[1:]:
        rec = {}
        for h, v in zip(headers, row):
            if h and v not in (None, ""):
                rec[h] = str(v).strip()
        if rec:
            agents.append(rec)
    return agents


def render_concept(
    rec: dict, scheme_iri: str, scheme_prefix: str = ":"
) -> tuple[str, str, bool]:
    """Render one concept block. Returns (ttl_block, concept_iri, is_top_concept)."""
    original_iri = normalize_iri(str(rec["conceptIRI"]))
    slug = iri_slug(original_iri)
    concept_iri = f"{scheme_iri}/{slug}"
    qname = f"{scheme_prefix}{slug}"

    lines = [f"{qname}"]
    lines.append("    a skos:Concept ;")
    lines.append("    rdfs:isDefinedBy cs: ;")
    lines.append("    skos:inScheme cs: ;")

    if rec.get("prefLabel"):
        lines.append(f"    skos:prefLabel {ttl_string(rec['prefLabel'])} ;")
    if rec.get("notation"):
        lines.append(f"    skos:notation {ttl_plain(rec['notation'])} ;")
    if rec.get("altLabel"):
        labels = split_list(rec["altLabel"])
        if len(labels) == 1:
            lines.append(f"    skos:altLabel {ttl_string(labels[0])} ;")
        elif labels:
            joined = " ,\n        ".join(ttl_string(l) for l in labels)
            lines.append(f"    skos:altLabel\n        {joined} ;")
    if rec.get("definition"):
        lines.append(f"    skos:definition {ttl_string(rec['definition'])} ;")
    if rec.get("provenance"):
        lines.append(f"    dcterms:provenance {ttl_string(rec['provenance'])} ;")
    if rec.get("historyNote"):
        lines.append(f"    skos:historyNote {ttl_string(rec['historyNote'])} ;")
    if rec.get("editorialNote"):
        lines.append(f"    skos:editorialNote {ttl_string(rec['editorialNote'])} ;")
    if rec.get("changeNote"):
        lines.append(f"    skos:changeNote {ttl_string(rec['changeNote'])} ;")
    if rec.get("usageNote"):
        lines.append(f"    vann:usageNote {ttl_string(rec['usageNote'])} ;")
    if rec.get("registryStatus"):
        lines.append(f"    reg:status <{rec['registryStatus'].strip()}> ;")

    is_top = bool(rec.get("isTopConceptOf"))
    if is_top:
        lines.append("    skos:topConceptOf cs: ;")

    lines.append(".")
    return "\n".join(lines), concept_iri, is_top


def render_agent(a: dict) -> str | None:
    """Render a background-label-style agent block, if the agent has a URI."""
    uri = a.get("agentURI")
    if not uri:
        return None
    uri = uri.strip()
    atype = a.get("agentType") or "schema:Thing"
    # normalise schema:Person / schema:Organization shorthand
    if atype.lower().startswith("schema:"):
        atype = atype
    lines = [f"<{uri}>"]
    lines.append(f"    a {atype} ;")
    if a.get("agentName"):
        lines.append(f"    schema:name {ttl_plain(a['agentName'])} ;")
    if a.get("agentEmail"):
        lines.append(f"    schema:email {ttl_plain(a['agentEmail'])} ;")
    if a.get("affiliationURI"):
        lines.append(f"    schema:affiliation <{a['affiliationURI'].strip()}> ;")
    lines.append(".")
    return "\n".join(lines)


def render_ttl(xlsx_path: Path) -> tuple[str, str]:
    """Read an xlsx file and return (ttl_text, scheme_slug)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Resolve sheet names (the source files sometimes mis-spell "Namespeces").
    sheet_map = {s.lower(): s for s in wb.sheetnames}
    cs_sheet = sheet_map.get("concept scheme")
    concepts_sheet = sheet_map.get("concepts")
    agents_sheet = sheet_map.get("agents")

    cs_data = load_sheet_kv(wb[cs_sheet]) if cs_sheet else {}
    concepts = load_concepts(wb[concepts_sheet]) if concepts_sheet else []
    agents = load_agents(wb[agents_sheet]) if agents_sheet else []

    scheme_iri = normalize_iri(str(cs_data["vocabularyIRI"][0]))
    scheme_slug = iri_slug(scheme_iri)

    created = parse_date(cs_data.get("created", [None])[0])
    modified = parse_date(cs_data.get("modified", [None])[0])
    version_raw = cs_data.get("version", [None])[0]
    version = parse_date(version_raw) or (str(version_raw) if version_raw else None)

    pref_label = (cs_data.get("prefLabel") or cs_data.get("title") or [scheme_slug])[0]
    title = (cs_data.get("title") or cs_data.get("prefLabel") or [pref_label])[0]
    definition = (cs_data.get("definition") or cs_data.get("description") or [""])[0]
    description = (cs_data.get("description") or [""])[0]

    # Rendered concept blocks plus discovered top-concept set
    concept_blocks: list[str] = []
    top_concepts: list[str] = []
    for rec in concepts:
        block, concept_iri, is_top = render_concept(rec, scheme_iri, scheme_prefix=":")
        concept_blocks.append(block)
        if is_top:
            top_concepts.append(concept_iri)

    # Unique, preserve order
    seen = set()
    top_concepts = [c for c in top_concepts if not (c in seen or seen.add(c))]

    # Resolve primary agent IRIs for creator / publisher references.
    creator_iri = None
    publisher_iri = None
    creator_name = (cs_data.get("creator") or [None])[0]
    publisher_name = (cs_data.get("publisher") or [None])[0]
    for a in agents:
        roles = (a.get("hadRole") or "").lower()
        name = a.get("agentName") or ""
        uri = a.get("agentURI") or a.get("affiliationURI")
        if not uri:
            continue
        if "creator" in roles and creator_iri is None:
            creator_iri = uri.strip()
        if "publisher" in roles and publisher_iri is None:
            publisher_iri = uri.strip()
        if creator_name and name == creator_name and creator_iri is None:
            creator_iri = uri.strip()
        if publisher_name and name == publisher_name and publisher_iri is None:
            publisher_iri = uri.strip()
    # fallback: if we know an affiliation URI, use that for org-level creator/publisher
    if creator_iri is None or publisher_iri is None:
        for a in agents:
            if a.get("affiliationURI"):
                aff = a["affiliationURI"].strip()
                creator_iri = creator_iri or aff
                publisher_iri = publisher_iri or aff
                break

    # Build TTL
    out: list[str] = []
    out.append(f"PREFIX : <{scheme_iri}/>")
    out.append(f"PREFIX cs: <{scheme_iri}>")
    for p, ns in PREFIXES:
        out.append(f"PREFIX {p}: <{ns}>")
    out.append("")

    # ConceptScheme block
    cs_lines: list[str] = ["cs:"]
    cs_lines.append("    a")
    cs_lines.append("        owl:Ontology ,")
    cs_lines.append("        skos:ConceptScheme ;")
    cs_lines.append(f"    dcterms:title {ttl_string(title)} ;")
    cs_lines.append(f"    skos:prefLabel {ttl_string(pref_label)} ;")
    if definition:
        cs_lines.append(f"    skos:definition {ttl_string(definition)} ;")
    if description and description != definition:
        cs_lines.append(f"    dcterms:description {ttl_string(description)} ;")
    if cs_data.get("scopeNote"):
        cs_lines.append(f"    skos:scopeNote {ttl_string(cs_data['scopeNote'][0])} ;")
    if cs_data.get("usageNote"):
        cs_lines.append(f"    vann:usageNote {ttl_string(cs_data['usageNote'][0])} ;")
    if cs_data.get("historyNote"):
        cs_lines.append(f"    skos:historyNote {ttl_string(cs_data['historyNote'][0])} ;")
    if cs_data.get("note"):
        cs_lines.append(f"    skos:note {ttl_string(cs_data['note'][0])} ;")
    if cs_data.get("provenance"):
        cs_lines.append(f"    dcterms:provenance {ttl_string(cs_data['provenance'][0])} ;")
    # Keywords — emitted as literal tags
    if cs_data.get("keywords"):
        kws = split_list(cs_data["keywords"][0])
        if kws:
            joined = " ,\n        ".join(ttl_string(k) for k in kws)
            cs_lines.append(f"    schema:keywords\n        {joined} ;")
    if cs_data.get("citation"):
        cs_lines.append(f"    schema:citation {ttl_string(cs_data['citation'][0])} ;")
    # Derived-from IRIs (keep literal strings so we don't lose non-IRI sources)
    if cs_data.get("derivedFrom"):
        df_items = split_list(cs_data["derivedFrom"][0])
        iri_items = [d for d in df_items if d.startswith("http")]
        for d in iri_items:
            cs_lines.append(f"    prov:wasDerivedFrom <{d}> ;")
    if cs_data.get("status"):
        cs_lines.append(f"    reg:status <{cs_data['status'][0].strip()}> ;")
    if cs_data.get("license"):
        cs_lines.append(f"    schema:license <{cs_data['license'][0].strip()}> ;")
    if creator_iri:
        cs_lines.append(f"    schema:creator <{creator_iri}> ;")
    elif creator_name:
        cs_lines.append(f"    schema:creator {ttl_plain(creator_name)} ;")
    if publisher_iri:
        cs_lines.append(f"    schema:publisher <{publisher_iri}> ;")
    elif publisher_name:
        cs_lines.append(f"    schema:publisher {ttl_plain(publisher_name)} ;")
    if created:
        cs_lines.append(f'    schema:dateCreated "{created}"^^xsd:date ;')
    if modified:
        cs_lines.append(f'    schema:dateModified "{modified}"^^xsd:date ;')
    if version:
        cs_lines.append(f"    schema:version {ttl_plain(version)} ;")
    if top_concepts:
        rendered_tops = []
        for iri in top_concepts:
            slug = iri_slug(iri)
            rendered_tops.append(f":{slug}")
        joined = " ,\n        ".join(rendered_tops)
        cs_lines.append(f"    skos:hasTopConcept\n        {joined} ;")
    cs_lines.append(".")
    out.append("\n".join(cs_lines))
    out.append("")

    # Agent blocks
    for a in agents:
        block = render_agent(a)
        if block:
            out.append(block)
            out.append("")

    # Concept blocks
    for block in concept_blocks:
        out.append(block)
        out.append("")

    return "\n".join(out).rstrip() + "\n", scheme_slug


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for xlsx in sorted(XLSX_DIR.glob("*.xlsx")):
        ttl, slug = render_ttl(xlsx)
        out_path = OUT_DIR / f"{slug}.ttl"
        out_path.write_text(ttl, encoding="utf-8")
        print(f"wrote {out_path.relative_to(ROOT)}  ({len(ttl)} bytes)")


if __name__ == "__main__":
    main()
